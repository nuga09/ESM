import pandas as pd
import os


def check_renewable_input(scenario_definition, paths):
    # check the coupled renewable energies
    if scenario_definition["renewables"] is not None:
        for tech in scenario_definition["renewables"].keys():
            if tech not in ["onshore", "offshore", "openfield_pv", "rooftop_pv"]:
                raise ValueError(
                    "The scenario definition contains an unknown technology" +
                    f"' {tech}'.Currently only ['onshore', 'offshore', " +
                    "'openfield_pv', 'rooftop_pv']are known")
            check_renewable_scenario(
                paths["nestor_ee_DB"], tech,
                scenario_definition["renewables"][tech]["nestor_ee_case"])


def check_renewable_scenario(path, tech, scenario):
    historical = pd.read_csv(
        os.path.join(path, tech, scenario, "historical.csv"), index_col=0)
    potentials = pd.read_csv(os.path.join(
        path, tech, scenario, "potentials.csv"), index_col=0)
    time_series = pd.read_csv(os.path.join(
        path, tech, scenario, "timeseries.csv"), index_col=0)

    # check if all components match
    ts_stock_names = [x.replace("_stock", "")
                      for x in time_series.columns if "_stock" in x]
    hist_names = [x for x in historical.columns]
    if sorted(ts_stock_names) != sorted(hist_names):
        print("Time-Series columns")
        print(ts_stock_names)
        print("Historical data columns:")
        print(hist_names)
        raise ValueError(
            "Mismatch in files. Time-Series and historical have different " +
            f"components for '{tech}' with '{scenario}'.")
    if sorted(list(time_series.columns)) != sorted(list(potentials["profile"].unique())):
        raise ValueError(
            "Mismatch between profiles of components in the potential.csv" +
            f" and the columns in timeseries.csv  for '{tech}' with '{scenario}'.")

    # check if time-series are between 0 and 1
    for component in time_series.columns:
        if (time_series[component] > 1).any():
            raise ValueError(
                f"Some time-series for '{tech}' with '{scenario}' are higher than 1")
        if (time_series[component] < 0).any():
            raise ValueError(
                f"Some time-series for '{tech}' with '{scenario}' are lower than 0")

    # check ub and lb stuff
    for component in historical.columns:
        # check that potentials are same or higher than historical in 2020
        if round(historical.loc["2020", component], 4) > round(potentials.loc[component, "ub"], 4):
            raise ValueError(
                f"Component '{component}' has a historical capacity of " +
                f"'{historical.loc['2020',component]}' but an upper bound of '" +
                f"{potentials.loc[component,'ub']}'. The ub must be equal or" +
                " higher than the historical capacity")
        if any(historical[component]<0):
            raise ValueError(
                "There are negative capacities for historical data for "+
                f"'{tech}' with '{scenario}'")
        # check ub and lb
        if potentials.loc[component, "ub"] < 0:
            raise ValueError(
                f"The upper bound for '{component}' is smaller than 0 " +
                "in potentials.csv")
        if potentials.loc[component, "lb"] < 0:
            raise ValueError(
                f"The lower bound for '{component}' is smaller than 0 " +
                "in potentials.csv")


def check_scenario_definition(json, local_mainpath):
    # datalocation
    if type(json["datalocation"]) is not str:
        raise TypeError("datalocation should be type str")
    if json["datalocation"] not in ['local', 'CAESAR']:
        raise ValueError("datalocation should either be 'local' or 'CAESAR'.")

    if json["datalocation"] == 'local':
        input_path = os.path.join(local_mainpath, "input_data")
        subfolders = \
            [f.name for f in os.scandir(input_path) if f.is_dir()]
        folder_names = ["forced_decommissioning", "historical_data",
                        "parameter", "template_evaluation_files", "timeseries"]
        _missing_list = [x for x in folder_names if x not in subfolders]
        if len(_missing_list) > 0:
            print("Folder missing in input_data: {}".format(_missing_list))
            raise ValueError("Download the Nestor-Data from CAESAR!")

    # QP
    if json["QP"] != False and json["QP"] != True:
        raise ValueError("QP in json should be true or false")

    # years
    if type(json["startyear"]) is not int:
        raise TypeError("startyear should be type int")
    if type(json["targetyear"]) is not int:
        raise TypeError("targetyear should be type int")
    if type(json["interval"]) is not int:
        raise TypeError("interval should be type int")
    if json["referenceyear"] != 2020:
        raise ValueError("Currently only referenceyear 2020 is supported")
    if json["targetyear"] > 2050 or json["targetyear"] < 2025:
        raise ValueError(
            "Currently only targetyears from 2020 to 2050 are supported")
    if json["interval"] != 5:
        raise ValueError("Currently only interval of 5 years is supported")
    if json["startyear"] <= json["referenceyear"]:
        raise ValueError("Referenceyear must be one interval before startyear")

    # exportflowresults
    if json["exportFlowResults"] != True and json["exportFlowResults"] != False:
        raise ValueError("exportFlowResults in json should be true or false")
    # max Refurbishment rate
    if json["maxRefurbRate"] > 3 or json["maxRefurbRate"] < 1:
        raise Warning("Check for realistic maximum refurbishment rate")
    # typical days
    if json["typdays"] is not None:
        if (int(json["typdays"]) > 365) or (int(json["typdays"]) < 0):
            raise ValueError(
                "Typical day should either be 'null' or value " +
                "between 1 and 365")
    # wacc
    if float(json["WACC"]) > 0.15 or float(json["WACC"]) < 0:
        raise Warning(
            "Realistic range of WACC in json should be between 0 and 0.15")

    # SCurve parameters
    if type(json["sCurveParam"]) is not list and len(json["sCurveParam"]) != 3:
        raise ValueError("sCurveParam should be list of three values")
    if json["sCurveParam"][0] < 1 or json["sCurveParam"][0] > 1.2:
        raise ValueError("sCurveParam in json should be between 1 and 1.2")
    if json["sCurveParam"][1] < 0 or json["sCurveParam"][1] > 0.5:
        raise ValueError("Doublecheck second value in scurve Param in json")
    if json["sCurveParam"][2] < 100 or json["sCurveParam"][2] > 150:
        raise ValueError("Doublecheck third value in scurve Param in json")

    # string_identifiers
    if type(json["string_identifiers"]) is not dict:
        raise TypeError("string_identifiers in json should be a dict")
    mandatory_string_identifiers = [
        "controlunit", "geothermalHeatPumps", "TSource", "TSink", "CO2Sinks", "onshore", "offshore", "rooftop_pv", "openfield_pv", "AEDemand", "LightingN-AEDemand", "IndustryDemand", "BatteryElectricCar",
        "BatteryElectricCar_Storage", "Sanierungspaket1", "building_exclusion_list", "building_standards", "building_filter_for_helper_components", "refurbishmentpackages", "building_categories"]
    json_str_identifiers = [x for x in json["string_identifiers"].keys()]
    if sorted(mandatory_string_identifiers) != sorted(json_str_identifiers):
        missing = [
            x for x in mandatory_string_identifiers if x not in json_str_identifiers]
        raise ValueError("Json string_identifiers missing: "+str(missing))

    # yearly expansion of renewables
    if json["maxYearlyExpansionGW_Onshore"] > 15:
        raise Warning("Check if maxYearlyExpansionGW_Onshore is realistic")
    if json["maxYearlyExpansionGW_Offshore"] > 10:
        raise Warning("Check if maxYearlyExpansionGW_Offshore is realistic")
    if json["maxYearlyExpansionGW_OFPV"] > 15:
        raise Warning("Check if maxYearlyExpansionGW_OFPV is realistic")
    if json["maxYearlyExpansionGW_RTPV"] > 15:
        raise Warning("Check if maxYearlyExpansionGW_RTPV is realistic")
    if json["considerMaxYearlyExpansionLimit"] != False and json["considerMaxYearlyExpansionLimit"] != True:
        raise TypeError(
            "AddResExpansionlimit in json should be true or false")
    if json["considerMaxYearlyExpansionLimit"] is True:
        import FINE as fn
        esM = fn.EnergySystemModel(locations=set(["Germany"]),
                                   commodities=set(["light", "heat"]),
                                   numberOfTimeSteps=8760,
                                   commodityUnitsDict={
                                       "light": "light", "heat": "heat"},
                                   hoursPerTimeStep=1, costUnit='1e9 Euro',
                                   lengthUnit='km', verboseLogLevel=0)

        try:
            esM.add(
                fn.Source(
                    esM=esM,
                    name="test",
                    commodity="light",
                    hasCapacityVariable=True,
                    operationRateMax=pd.Series([1] * 8760),
                    capacityFix=None,
                    capacityMax=1,
                    capacityMin=0,
                    investPerCapacity=100,
                    opexPerCapacity=0.02,
                    opexPerOperation=0.02,
                    commodityCost=1,
                    commodityRevenue=0,
                    economicLifetime=20,
                    technicalLifetime=20,
                    sharedCapacityExpansionMax=50,
                    sharedCapacityExpansionMin=0,
                    sharedExpansionID="test"))

        except Exception as e:
            print(e)
            raise ValueError("If you want to limit the expansion of renewables, " +
                             "you must use a FINE branch with 'sharedExpansionMin' and " +
                             "'sharedExpansionMax'.")

    # reasonable value for usable bev storage capacity
    if float(json["UsableBEVStorageCapacity_per_km"]) > 5 or float(json["UsableBEVStorageCapacity_per_km"]) < 0:
        raise Warning(
            "UsableBEVStorageCapacity_per_km in json has no realistic value")

    # check for solver
    if json["solver"] not in ["gurobi", "glpk"]:
        raise ValueError("Solver must be either 'gurobi' or 'glpk'")

    if json["solver"] == "glpk" and json["Run"] != "Mini-System":
        raise ValueError(
            "Solver 'glpk' only usable for testing the minisystem in the " +
            "CI-pipeline in jugit. Please use gurobi")


def check_input_data(opt_json, paths, scenario_definition):
    # get list of unrestrited components
    raw_transformers = pd.read_excel(
        paths["parameterfile"], sheet_name='Transformers', index_col="name")
    raw_sources = pd.read_excel(
        paths["parameterfile"], sheet_name='Sources', index_col=0)
    raw_heatpumps = pd.read_excel(
        paths["parameterfile"], sheet_name='HeatPumps', index_col=0)
    raw_storages = pd.read_excel(
        paths["parameterfile"], sheet_name='Storages', index_col=0)
    raw_sinks = pd.read_excel(
        paths["parameterfile"], sheet_name='Sinks', index_col=0)
    raw_hubs = pd.read_excel(
        paths["parameterfile"], sheet_name='Hubs', index_col=0)
    connections = pd.read_excel(
        paths["parameterfile"], "Connectors")
    all_components = list(raw_transformers.index)+list(raw_sources.index) +\
        list(raw_heatpumps.index) + list(raw_hubs.index) +\
        list(raw_storages.index)+list(raw_sinks.index)

    unrestricted_components = pd.concat([
        raw_transformers.loc[raw_transformers["unrestrict"] == True],
        raw_sources.loc[raw_sources["unrestrict"] == True],
        raw_storages.loc[raw_storages["unrestrict"] == True],
        raw_sinks.loc[raw_sinks["unrestrict"] == True],
        raw_hubs.loc[raw_hubs["unrestrict"] == True]],
        axis=0).index

    check_optimization_definition(opt_json)
    check_parameter_file(paths["parameterfile"])
    check_stock_hist_data(
        paths["parameterfile"],
        paths["historicaldata"])
    check_if_components_with_lb_have_historical_capacity(
        paths["parameterfile"],
        paths["historicaldata"])
    compare_unrestriced_components_and_historical_data(
        paths["historicaldata"], unrestricted_components)
    check_forced_decommissioning_data(
        paths["forceddecommissioningdata"],
        paths["historicaldata"],
        paths["parameterfile"],
        scenario_definition["targetyear"])
    check_if_unrestriced_have_ub(
        unrestricted_components, paths["parameterfile"])
    check_if_every_component_is_connected(
        all_components, connections)
    check_if_every_entry_in_connections_is_component(
        all_components, connections)
    check_if_stock_have_changing_parameters_or_costscale(
        raw_transformers, raw_sources, raw_heatpumps, raw_storages, raw_sinks, raw_hubs)


def check_if_every_component_is_connected(all_components, raw_connections):
    all_connections = list(
        raw_connections["input"])+list(raw_connections["output"])
    unconnected_components = []
    for component in all_components:
        if component not in all_connections:
            unconnected_components.append(component)
    if len(unconnected_components) > 0:
        raise ValueError("Some components are not connected: " +
                         "{}".format(unconnected_components))


def check_if_every_entry_in_connections_is_component(all_components, raw_connections):
    missing_components_in_input_connection = []
    for input in raw_connections["input"]:
        if input not in all_components:
            missing_components_in_input_connection.append(input)
    missing_components_in_output_connection = []
    for output in raw_connections["output"]:
        if output not in all_components:
            missing_components_in_output_connection.append(output)
    if len(missing_components_in_input_connection) > 0 or len(missing_components_in_output_connection) > 0:
        raise ValueError(
            "Components referenced in sheet 'Connections' do not exist as component in the other sheets: " +
            "{} {}".format(missing_components_in_input_connection, missing_components_in_output_connection))


def check_if_stock_have_changing_parameters_or_costscale(
        raw_transformers, raw_sources, raw_heatpumps, raw_storages, raw_sinks, raw_hubs):
    def check_stocks(df, cols):
        error = False
        stocks = [x for x in df.index if "_stock" in x]
        _df = df.loc[stocks]
        for param, param_list in cols.items():
            # check if any stock has a costscale
            if param == "costscale":
                _stocks_with_costscale = []
                for cost_scale_col in param_list:
                    stock_with_costscale = _df.loc[_df[cost_scale_col] > 0]
                    _stocks_with_costscale.extend(
                        stock_with_costscale.index.tolist())
                if len(_stocks_with_costscale) > 0:
                    print(
                        f"The following stocks wrongly have a costscale: {_stocks_with_costscale}")
                    error = True

            # check if all stock components are same over transformation time
            for idx, rows in _df.iterrows():
                if len(rows[param_list].unique()) > 1:
                    print(
                        f"Stock Component '{idx}' has different parameters for '{param}'")
                    error = True
        if error == True:
            raise ValueError("Check log above.")

    source_cols = {"capex": ["capex20", "capex30", "capex40", "capex50"],
                   "costscale": [
        "cost_scale2020", "cost_scale2050"]}
    transformers_cols = {"capex": ["capex20", "capex30", "capex40", "capex50"],
                         "costscale": [
        "cost_scale2020", "cost_scale2050"]}
    sink_cols = {"costscale": ["cost_scale2020"]}
    storage_cols = {"capex": ["capex20", "capex30", "capex40", "capex50"],
                    "costscale": [
        "cost_scale2020", "cost_scale2050"]}
    heatpump_cols = {"capex": ["capex20", "capex30", "capex40", "capex50"],
                     "costscale": [
        "cost_scale2020", "cost_scale2050"]}
    check_stocks(raw_sources, source_cols)
    check_stocks(raw_transformers, transformers_cols)
    check_stocks(raw_heatpumps, heatpump_cols)
    check_stocks(raw_storages, storage_cols)
    check_stocks(raw_sinks, sink_cols)


def check_optimization_definition(json):
    if type(json["Threads"]) is not int or json["Threads"] > 40:
        raise ValueError("Please give realistic number of threads in json.")
    if type(json["Big_M"]) is not int or json["Big_M"] > 300 or json["Big_M"] < 100:
        raise Warning("Please give realistic number of BigM in json.")
    if type(json["Numeric_Focus"]) is not int or json["Numeric_Focus"] > 3 or json["Numeric_Focus"] < 0:
        raise ValueError("Numeric Focus in json should be 0,1,2 or 3")
    if type(json["Crossover"]) is not int or json["Crossover"] > 4 or json["Crossover"] < 0:
        raise ValueError("Crossover in json should be 0,1,2, 3 or 4")
    if type(json["NodeMethod"]) is not int or json["NodeMethod"] > 2 or json["NodeMethod"] < 0:
        raise ValueError("NodeMethod in json should be int and 0,1 or 2")
    if type(json["Method"]) is not int or json["Method"] > 5 or json["Method"] < -1:
        raise ValueError("Method in json should be int and from -1 to 5")
    if json["getDual"] != False and json["getDual"] != True:
        raise ValueError("getDual in json should be true or false")
    if type(json["BarHomogeneous"]) is not int or json["BarHomogeneous"] > 1 or json["BarHomogeneous"] < -1:
        raise ValueError("BarHomogeneous in json should be int from -1 to 1")
    if type(json["Opt_Tolerance"]) is not int or json["Opt_Tolerance"] > 10 or json["Opt_Tolerance"] < 1:
        raise Warning(
            "A realistic Opt_Tolerance in json should be int from 1 and 10")
    if json["tee"] != False and json["tee"] != True:
        raise ValueError("tee in json should be true or false")
    if json["Backcasting"] != False and json["Backcasting"] != True:
        raise ValueError("Backcasting in json should be true or false")
    if type(json["Max_Expansion"]) is not float or json["Max_Expansion"] > 0.3 or json["Max_Expansion"] < 0.1:
        raise Warning(
            "A realistic Max_Expansion in json should be float between 0.1 and 0.3")
    if type(json["Max_Decomm"]) is not float or json["Max_Decomm"] > 0.6 or json["Max_Decomm"] < 0.4:
        raise Warning(
            "A realistic Max_Decomm in json should be float between 0.4 and 0.6")
    if type(json["Max_Dev_Factor"]) is not float or json["Max_Dev_Factor"] > 1.2 or json["Max_Dev_Factor"] < 1:
        raise Warning(
            "A realistic Max_Dev_Factor in json should be float between 1 and 1.2")
    if type(json["Max_Add_Dev"]) is not float or json["Max_Add_Dev"] > 0.1 or json["Max_Add_Dev"] < 0:
        raise Warning(
            "A realistic Max_Add_Dev in json should be float between 0 and 0.1")
    if type(json["First_Year_Cor"]) is not float or json["First_Year_Cor"] > 1.2 or json["First_Year_Cor"] < 1:
        raise Warning(
            "A realistic First_Year_Cor in json should be float between 1 and 1.2")


def check_parameter_file(param_file_path):
    hubs = pd.read_excel(param_file_path, "Hubs", index_col="name")
    transformer_df = pd.read_excel(
        param_file_path, "Transformers", index_col="name")
    connections = pd.read_excel(param_file_path, "Connectors")
    sources = pd.read_excel(param_file_path, "Sources", index_col="name")
    sinks = pd.read_excel(param_file_path, "Sinks", index_col="name")
    storages = pd.read_excel(param_file_path, "Storages", index_col="name")
    heatpumps = pd.read_excel(
        param_file_path, "HeatPumps", index_col="name")

    # check functions
    check_for_duplicate_components(
        transformer_df, sources, sinks, storages, heatpumps)
    check_for_opex_for_hubs(hubs=hubs)
    check_input_output_energytypes_of_transformer(
        transformer_df, connections, hubs, sources)
    check_hub_after_source_with_co2footprint(
        hubs=hubs, sources=sources, connections=connections)
    check_hub_before_sink_with_co2footprint(hubs, sinks, connections)
    check_dimenergy_and_energytype_of_transformer(transformer_df)
    check_if_stocks_have_ub_greater_zero(
        transformer_df, sources, sinks, storages, heatpumps)
    check_if_any_cell_is_formular(param_file_path)


def check_if_any_cell_is_formular(param_file_path):
    from openpyxl import load_workbook
    book = load_workbook(param_file_path)
    error = False
    for sheet_name in book.sheetnames:
        sheet = book[sheet_name]
        for row in sheet.iter_rows(min_row=1):
            for cell in row:
                if cell.data_type == "f":
                    print("Cell with formula in sheet: {}".format(sheet_name))
                    print("Row: {}".format(row))
                    print("Value: {}".format(cell.value))
                    error = True
    if error:
        raise ValueError("Remove cells with formula in Parameter-File!")


def check_for_duplicate_components(transformer, sources, sinks, storages, heatpumps):
    # parameter file with ubs
    raw_ub = pd.concat([
        transformer["ub"],
        sources["ub"],
        heatpumps["ub"],
        storages["ub"],
        sinks["ub"]], axis=0).round(5)

    if any(raw_ub.index.duplicated()):
        duplications = raw_ub.index[raw_ub.index.duplicated()].unique()

        for i in duplications:
            print(i)
            for df in [transformer, sources, sinks, storages, heatpumps]:
                if i in df.index:
                    print(df)
                    print(df.loc[i])
        raise ValueError(
            "Components are double defined in parameter-file:" +
            "'{}'".format(duplications))


def check_for_duplicate_components(transformer, sources, sinks, storages, heatpumps):
    # parameter file with ubs
    raw_ub = pd.concat([
        transformer["ub"],
        sources["ub"],
        heatpumps["ub"],
        storages["ub"],
        sinks["ub"]], axis=0).round(5)

    if any(raw_ub.index.duplicated()):
        duplications = raw_ub.index[raw_ub.index.duplicated()].unique()
        for i in duplications:
            print(i)
        raise ValueError(
            "Components are double defined in parameter-file:" +
            "'{}'".format(duplications))


def check_dimenergy_and_energytype_of_transformer(transformer):
    different_energytypes = transformer.loc[transformer[
        "dimEnergyType"] != transformer["energytype"]].index
    if len(different_energytypes) > 0:
        raise ValueError("Transformers {} ".format(different_energytypes) +
                         "should not have different energytype and " +
                         "dimEnergyType in Parameterfile.")


def check_if_stocks_have_ub_greater_zero(transformer, sources, sinks,
                                         storages, heatpumps):
    def _check_stock_ub_greater_zero(df):
        stock_index = [x for x in df.index if "_stock" in x]
        stock_df = df.loc[stock_index]
        stock_with_ub_greater_zero = stock_df.loc[stock_df["ub"] != 0].index

        if len(stock_with_ub_greater_zero) > 0:
            raise ValueError(
                "Stocks '{}' ".format(stock_with_ub_greater_zero) +
                "should not have an upper bound in parameter file.")

    _check_stock_ub_greater_zero(transformer)
    _check_stock_ub_greater_zero(sources)
    _check_stock_ub_greater_zero(sinks)
    _check_stock_ub_greater_zero(storages)
    _check_stock_ub_greater_zero(heatpumps)


def check_input_output_energytypes_of_transformer(transformer_df, connections, hubs, sources):
    """Check that the input energy type and output energy type of transformer
    are different.

    Parameters
    ----------
    transformer_df : pd.DataFrame
    connections : pd.DataFrame
    hubs : pd.DataFrame
    """
    error = False
    for transformer, transformer_parm in transformer_df.iterrows():
        if "rSOFC" in transformer:
            continue
        dim_energy = transformer_parm["energytype"]
        output_connections = connections.loc[
            connections["input"] == transformer, "output"]
        input_connections = connections.loc[
            connections["output"] == transformer, "input"]

        input_energytypes = []
        for idx, comp in input_connections.iteritems():
            if comp in hubs.index:
                input_energytypes.append(hubs.loc[comp].energytype)
            elif comp in sources.index:
                input_energytypes.append(sources.loc[comp].energytype)
            else:
                print("Cannot match {} for {}".format(comp, transformer))

        output_energytypes = []
        for idx, comp in output_connections.iteritems():
            output_energytypes.append(hubs.loc[comp].energytype)

        intersect = [value for value in input_energytypes
                     if value in output_energytypes]
        if len(intersect) > 0:
            print(
                "Transformer {} has energytype {} as input and output".format(
                    transformer, intersect))
            error = True
    if error:
        raise ValueError(
            "Transformers with same input energy type as output energytype")


def check_for_opex_for_hubs(hubs):
    """Check that hubs do not have opex costs.

    Parameters
    ----------
    transformer_df : pd.DataFrame
    connections : pd.DataFrame
    hubs : pd.DataFrame
    """
    hubs_with_fix_opex = hubs.loc[hubs["opex_fix"] != 0]
    hubs_with_var_opex = hubs.loc[hubs["opex_var"] != 0]
    hubs_with_capex = hubs.loc[hubs["capex"] != 0]

    if len(hubs_with_fix_opex) > 0:
        raise ValueError(
            "Hubs ({}) with opex. ".format(hubs_with_fix_opex.index) +
            "Please use an additional transformer for including opex")

    if len(hubs_with_var_opex) > 0:
        raise ValueError(
            "Hubs ({}) with opex. ".format(hubs_with_var_opex.index)
            + "Please use an additional transformer for including opex")

    if len(hubs_with_capex) > 0:
        raise ValueError(
            "Hubs ({}) with capex. ".format(hubs_with_var_opex.index)
            + "Hubs should not havee capex.")


def check_hub_before_sink_with_co2footprint(hubs, sinks, connections):
    """Check that the hub before a sink with CO2footprint only has one output connection (to the sink).

    Parameters
    ----------
    hubs : pd.DataFrame
    sinks : pd.DataFrame
    """
    error = False
    for hub, hub_param in hubs.iterrows():
        hub_connections = connections.loc[connections["input"] == hub]
        hub_sink_connections = [
            x for x in hub_connections["output"] if x in sinks.index]
        hub_sink_connections = [
            x for x in hub_sink_connections if not "TSink" in x]
        if len(hub_sink_connections) > 1:
            if any(sinks.loc[hub_sink_connections, "CO2footprint"] > 0):
                print("Previous hub '{}' of sink with a CO2 footprint with too many output connections ({}). Hub before Sink with CO2footprint should just have one connection.".format(
                    hub, hub_sink_connections))
                error = True
    if error:
        raise ValueError(
            "Some Hubs before Sink with CO2footprint  with too many connections. Check log.")


def check_hub_after_source_with_co2footprint(hubs, sources, connections):
    """Check that the hub after a source with CO2footprint only has one input connection (to the source).

    Parameters
    ----------
    hubs : pd.DataFrame
    sinks : pd.DataFrame
    """
    error = False
    for hub, hub_param in hubs.iterrows():
        hub_connections = connections.loc[connections["output"] == hub]
        hub_sources_connections = [
            x for x in hub_connections["input"] if x in sources.index]
        hub_sources_connections = [
            x for x in hub_sources_connections if not "TSource" in x]
        if len(hub_sources_connections) > 1:
            if any(sources.loc[hub_sources_connections, "CO2footprint"] > 0):
                print("Hub '{}' of source with a CO2 footprint with too many output connections ({}). Hub after Source with CO2footprint should just have one connection.".format(
                    hub, hub_sources_connections))
                error = True
    if error:
        raise ValueError(
            "Some Hubs after Source with CO2footprint with too many connections. Check log.")


def check_if_components_with_lb_have_historical_capacity(parameter_file_path, historical_cap_path):
    # 1. read data
    # historical data
    historical_data = pd.read_excel(historical_cap_path, index_col=0)

    # parameter file with ubs
    transformers = pd.read_excel(
        parameter_file_path, sheet_name='Transformers', index_col="name")
    raw_sources = pd.read_excel(
        parameter_file_path, sheet_name='Sources', index_col=0)
    raw_heatpumps = pd.read_excel(
        parameter_file_path, sheet_name='HeatPumps', index_col=0)
    raw_storages = pd.read_excel(
        parameter_file_path, sheet_name='Storages', index_col=0)
    raw_sinks = pd.read_excel(
        parameter_file_path, sheet_name='Sinks', index_col=0)

    raw_lb = pd.concat([
        transformers["lb"],
        raw_sources["lb"],
        raw_heatpumps["lb"],
        raw_storages["lb"],
        raw_sinks["lb"]], axis=0).round(5)

    # 2. check if every component in lb has historical data
    components_with_lb = raw_lb.loc[raw_lb > 0].index
    components_with_lb_without_historical_data = \
        [x for x in components_with_lb if x not in historical_data.columns]
    if len(components_with_lb_without_historical_data) > 0:
        print(
            "\nWarning: Some components have lb but no historical data: " +
            "{}".format(components_with_lb_without_historical_data))


def check_stock_hist_data(parameterfile_path, histData_path):
    """Check if stocks of parameter file and historical data file match.

    Parameters
    ----------
    parameterfile_path : str
    histData_path : str
    """

    # get list of stock components in parameter file
    stocks_parameterfile = []
    sheetnames = ['Sources', 'Sinks', 'Storages', 'Transformers', 'HeatPumps']
    for sheetname in sheetnames:
        parameterfile = pd.read_excel(
            parameterfile_path, sheet_name=sheetname, index_col=0)
        _componentenlist_stock = [
            x for x in parameterfile.index if "_stock" in x]
        stocks_parameterfile.extend(_componentenlist_stock)

    stocks_parameterfile = [x.replace("_stock", "")
                            for x in stocks_parameterfile]

    # get list of stock components in historical data
    hist_data = pd.read_excel(histData_path, index_col=0)
    stocks_historicaldata = hist_data.columns

    # compare lists and raise error if stocks of parameterfile and historical
    # data do not match
    # _additional_stocks_in_parameterfile=stocks_parameterfile.remove(stocks_historicaldata)

    _additional_stocks_in_parameterfile = [
        x for x in stocks_parameterfile if x not in stocks_historicaldata]
    if len(_additional_stocks_in_parameterfile) > 0:
        raise ValueError(
            "Stocks {} ".format(_additional_stocks_in_parameterfile)
            + "in Parameter-File but not in historical data.")

    _additional_stocks_in_historicaldata = [
        x for x in stocks_historicaldata if x not in stocks_parameterfile]
    if len(_additional_stocks_in_historicaldata) > 0:
        raise ValueError(
            "Stocks {} ".format(_additional_stocks_in_historicaldata)
            + "in historical data but not in parameter file.")


def compare_unrestriced_components_and_historical_data(historical_cap_path, unrestricted_components):
    historical_components = pd.read_excel(
        historical_cap_path, index_col=0).columns

    _missing_components = []

    for h_component in historical_components:
        if h_component in unrestricted_components:
            _missing_components.append(h_component)

    if len(_missing_components) > 0:
        print("\nWARNING: The following components have historical data but will be considered unrestricted: {} \n".format(
            _missing_components))


def check_forced_decommissioning_data(forced_decommissioning_path, historical_cap_path, parameter_file_path, targetyear):
    # 1. read data
    # historical data
    historical_data = pd.read_excel(historical_cap_path, index_col=0)

    # forced decommissioning
    forced_decommissioning = pd.read_excel(
        forced_decommissioning_path, index_col=0)
    unnamed_cols = [
        x for x in forced_decommissioning.columns if x.startswith("Unnamed: ")]
    forced_decommissioning = forced_decommissioning.drop(columns=unnamed_cols)
    # parameter file with ubs
    transformers = pd.read_excel(
        parameter_file_path, sheet_name='Transformers', index_col="name")
    raw_sources = pd.read_excel(
        parameter_file_path, sheet_name='Sources', index_col=0)
    raw_heatpumps = pd.read_excel(
        parameter_file_path, sheet_name='HeatPumps', index_col=0)
    raw_storages = pd.read_excel(
        parameter_file_path, sheet_name='Storages', index_col=0)
    raw_sinks = pd.read_excel(
        parameter_file_path, sheet_name='Sinks', index_col=0)

    raw_ub = pd.concat([
        transformers["ub"],
        raw_sources["ub"],
        raw_heatpumps["ub"],
        raw_storages["ub"],
        raw_sinks["ub"]], axis=0).round(5)

    # 2. check if forced_decommissioning components are in parameter file
    unmatched_components = [
        x for x in forced_decommissioning.columns if x not in raw_ub.index]
    if len(unmatched_components) > 0:
        raise ValueError(
            "Components in forced_decommissioning cannot be found in " +
            "parameter-file: {}".format(unmatched_components))

    # 3. check if forced_decomissioning <= historical data
    fd_years = forced_decommissioning.index
    for component in forced_decommissioning.columns:
        _forced_decommissioning = round(
            forced_decommissioning.loc[fd_years, component], 4)
        _historical = round(historical_data.loc[fd_years, component], 4)

        if any(_forced_decommissioning > _historical):
            for year in fd_years:
                if forced_decommissioning.loc[year, component] > historical_data.loc[year, component]:
                    print("Problem with forced decomissioning in year " + str(year))
                    print("forced_decommissioning.loc[year, component]")
                    print(forced_decommissioning.loc[year, component])
                    print("historical_data.loc[year, component]")
                    print(historical_data.loc[year, component])
            raise ValueError("Forced Decomissioning higher than historical data for " +
                             "'{}'".format(component))

    # 4. check if jede kompente mit ub < historicalcap[tj] auch forced
    #    decommsissioning hat
    components_with_ub_greater_historical_capacity_tj = [
        x for x in historical_data.columns if (raw_ub[x] < historical_data.loc[targetyear, x]).any()]

    _missing = [
        x for x in components_with_ub_greater_historical_capacity_tj if x not in forced_decommissioning.columns]

    if len(_missing) > 0:
        print("\n\n WARNING: List of components which need a forced decommissioning " +
              "pathway in the forced_commissioning.xlsx: {} \n\n".format(_missing))


def check_if_unrestriced_have_ub(unrestricted_components, parameter_file_path):
    # parameter file with ubs
    transformers = pd.read_excel(
        parameter_file_path, sheet_name='Transformers', index_col="name")
    raw_sources = pd.read_excel(
        parameter_file_path, sheet_name='Sources', index_col=0)
    raw_heatpumps = pd.read_excel(
        parameter_file_path, sheet_name='HeatPumps', index_col=0)
    raw_storages = pd.read_excel(
        parameter_file_path, sheet_name='Storages', index_col=0)
    raw_sinks = pd.read_excel(
        parameter_file_path, sheet_name='Sinks', index_col=0)

    raw_ub = pd.concat([
        transformers["ub"],
        raw_sources["ub"],
        raw_heatpumps["ub"],
        raw_storages["ub"],
        raw_sinks["ub"]], axis=0).round(5)

    _list = []
    for component in raw_ub.index:
        if component in unrestricted_components:
            if raw_ub[component] < 500:
                _list.append("{} with ub {}".format(
                    component, raw_ub[component]))

    if len(_list) > 0:
        print("WARNING: Unrestricted components with ub in parameter file: " +
              "{} \n".format(_list))
