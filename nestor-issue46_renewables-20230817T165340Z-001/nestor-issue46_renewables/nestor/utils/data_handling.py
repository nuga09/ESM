import os
import shutil
import datetime
from xml.dom import NotFoundErr
import openpyxl
from nestor.utils.coupling_renewables import PotentialUpdate


def get_raw_input_data_paths(local_mainpath, scenario_definition):
    if scenario_definition["datalocation"] == "CAESAR":
        mainpath = r"/storage/internal/data/nestor/"
    elif scenario_definition["datalocation"] == "local":
        mainpath = local_mainpath

    # Paths
    raw_input_data_paths = {}
    raw_input_data_paths["parameterfile"] =\
        os.path.join(mainpath, "input_data", "parameter",
                     scenario_definition["databasename"])
    raw_input_data_paths["historicaldata"] = \
        os.path.join(mainpath, "input_data", "historical_data",
                     scenario_definition["historicaldataname"])
    raw_input_data_paths["forceddecommissioningdata"] = \
        os.path.join(mainpath, "input_data", "forced_decommissioning",
                     scenario_definition["forceddecommissioningdataname"])

    profilepath = os.path.join(mainpath, "input_data", 'timeseries')
    raw_input_data_paths["heatload"] = \
        os.path.join(profilepath, scenario_definition["heatloaddataname"])

    raw_input_data_paths["input_profiles"] = \
        os.path.join(profilepath, scenario_definition["inputprofiledataname"])
    raw_input_data_paths["output_profiles"] = \
        os.path.join(profilepath, scenario_definition["outputprofiledataname"])

    raw_input_data_paths["nestor_ee_DB"] =\
        os.path.join(mainpath, "input_data", "renewables")

    raw_input_data_paths["template_evaluation"] = \
        os.path.join(mainpath, "input_data", "template_evaluation_files",
                     scenario_definition["evaluationfilename"])
    if not os.path.isfile(raw_input_data_paths["template_evaluation"]):
        raise NotFoundErr("Template of evaluation file not found ({})".format(
            raw_input_data_paths["template_evaluation"]))
    return raw_input_data_paths


def folder_creation(local_mainpath, raw_input_data_paths, scenario_definition,
                    scenariopath, GHGScenpath, optParapath):
    # result folder path
    respath = os.path.join(local_mainpath, "Results")

    # generate string/name for results folder
    now = datetime.datetime.now()
    resfoldername = (
        now.strftime("%Y-%m-%d %H_%M_%S") +
        ' QP_' + str(scenario_definition["QP"]) +
        ' TSA_' + str(scenario_definition["typdays"]) +
        ' Name_' + str(scenario_definition["Run"]))
    resfolderpath = os.path.join(respath, resfoldername)
    rawinput_folderpath = os.path.join(
        resfolderpath, "input", "raw_input_data")
    processedinput_folderpath = os.path.join(
        resfolderpath, "input", "processed_input_data")
    temppath = os.path.join(resfolderpath, 'temp')
    scendef_path = os.path.join(
        resfolderpath, 'input', 'scenario_definition')
    print(resfoldername, flush=True)

    # create Result Folder
    os.makedirs(resfolderpath)
    os.makedirs(temppath)
    os.makedirs(rawinput_folderpath)
    os.makedirs(processedinput_folderpath)
    os.makedirs(scendef_path)

    raw_input_paths_resfolder = {}
    folder_names = ["forced_decomissioning", "historical_data", "parameter",
                    "template_evaluation_files", "timeseries", "renewables"]
    for i in folder_names:
        raw_input_paths_resfolder[i] = os.path.join(rawinput_folderpath, i)
        os.makedirs(raw_input_paths_resfolder[i])

    ###
    paths = {}
    paths["temppath"] = temppath
    paths["resfolderpath"] = resfolderpath

    # 1.1. copy scenario definition
    shutil.copy(scenariopath, scendef_path)
    shutil.copy(optParapath, scendef_path)
    shutil.copy(GHGScenpath, scendef_path)

    # 1.2. Forced Decomissioning
    shutil.copy(
        raw_input_data_paths["forceddecommissioningdata"],
        raw_input_paths_resfolder["forced_decomissioning"])
    paths["forceddecommissioningdata"] = \
        os.path.join(processedinput_folderpath,
                     scenario_definition["forceddecommissioningdataname"])
    shutil.copy(
        raw_input_data_paths["forceddecommissioningdata"],
        processedinput_folderpath)

    # 1.3 Heat Load
    shutil.copy(raw_input_data_paths["heatload"],
                raw_input_paths_resfolder["timeseries"])
    paths["heatload"] = os.path.join(
        processedinput_folderpath, scenario_definition["heatloaddataname"])
    shutil.copy(raw_input_data_paths["heatload"],
                processedinput_folderpath)

    # 1.4 Output profiles
    shutil.copy(raw_input_data_paths["output_profiles"],
                raw_input_paths_resfolder["timeseries"])
    paths["output_profiles"] = os.path.join(
        processedinput_folderpath, scenario_definition["outputprofiledataname"])
    shutil.copy(raw_input_data_paths["output_profiles"],
                processedinput_folderpath)

    # 1.5 Parameterfile - only raw data
    shutil.copy(raw_input_data_paths["parameterfile"],
                os.path.join(raw_input_paths_resfolder["parameter"]))

    # 1.6 Historical data - only raw data
    shutil.copy(raw_input_data_paths["historicaldata"],
                raw_input_paths_resfolder["historical_data"])

    # 1.7 input profiles - only raw data
    shutil.copy(raw_input_data_paths["input_profiles"],
                raw_input_paths_resfolder["timeseries"])

    # 1.8 potentials - raw
    for renewable_tech, renewable_info in scenario_definition["renewables"].items():
        _re_old_path = os.path.join(
            raw_input_data_paths["nestor_ee_DB"],
            renewable_tech, renewable_info["nestor_ee_case"])
        _re_new_path = os.path.join(
            raw_input_paths_resfolder["renewables"], renewable_tech,
            renewable_info["nestor_ee_case"])
        # os.makedirs(_re_new_path)
        shutil.copytree(_re_old_path, _re_new_path)

    # PROCESS PARAMETER, HISTORICAL DATA AND INPUT PROFILES
    # initialize Potential coupling
    _path = os.path.join(
        processedinput_folderpath, scenario_definition["databasename"])
    potential_data = PotentialUpdate(
        renewable_definition=scenario_definition["renewables"],
        nestor_ee_path=raw_input_data_paths["nestor_ee_DB"],
        parameter_file_path=raw_input_data_paths["parameterfile"])
    # 2.1 Pameterfile processed
    paths["parameterfile"] = potential_data.update_parameter_files(
        targetpath=processedinput_folderpath
    )
    # 2.2 Input profiles processed
    paths["input_profiles"] = potential_data.update_timeseries(
        raw_input_data_paths["input_profiles"],
        targetpath=processedinput_folderpath)
    # 2.3 historical data processed
    paths["historicaldata"] = potential_data.update_historical_data(
        raw_input_data_paths["historicaldata"],
        targetpath=processedinput_folderpath)
    return paths


def create_evaluation_file(resultfolderpath, resultfilepath,
                           templateEvaluationPath, scenario_definition):
    wb_update = openpyxl.load_workbook(resultfilepath)
    wb_evaluation = openpyxl.load_workbook(templateEvaluationPath)
    for sheet_page in wb_update.sheetnames:
        if sheet_page == "costsLP":
            continue
        ws1 = wb_update[sheet_page]
        ws2 = wb_evaluation[sheet_page]
        # calculate total number of rows and columns in source excel file
        mr = ws1.max_row
        mc = ws1.max_column
        # read and write the cell values from result file to evaluation file
        for year in range(1, mr + 1):
            for j in range(1, mc + 1):
                c = ws1.cell(row=year, column=j)
                ws2.cell(row=year, column=j).value = c.value
    wb_evaluation.save(
        resultfolderpath +
        '/Evaluation_Results_{}.xlsx'.format(scenario_definition["Run"]))
